import streamlit as st, gspread, pandas as pd, calendar, io
from datetime import date
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

st.set_page_config(page_title="GAJI V16 XLSX", layout="wide", page_icon="💰")
st.markdown("<h2>💰 GAJI V16 - GENERATE XLSX</h2>", unsafe_allow_html=True)

@st.cache_resource
def connect_gsheet():
    scope = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scope)
    client = gspread.authorize(creds)
    sh = client.open("REKAP")
    return sh.worksheet("REKAP ABSENSI"), sh.worksheet("DATABASE KARYAWAN")
ws_absen, ws_db = connect_gsheet()

@st.cache_data(ttl=60)
def load_data():
    db = pd.DataFrame(ws_db.get_all_records())
    db['ID KARYAWAN'] = db['ID KARYAWAN'].astype(str).str.zfill(8)
    vals = ws_absen.get_all_values()
    HEADER = ['ID KARYAWAN','NAMA KARYAWAN','TANGGAL','JAM MASUK','JAM PULANG','JAM KERJA','JAM LEMBUR','LEMBUR 1.5','LEMBUR 2.0','SHIFT','KETERANGAN','STATUS','UANG SHIFT']
    if len(vals)>1:
        data = [r[:13] for r in vals[1:]]
        absen = pd.DataFrame(data, columns=HEADER) if data else pd.DataFrame(columns=HEADER)
    else:
        absen = pd.DataFrame(columns=HEADER)
    if not absen.empty:
        absen['ID KARYAWAN'] = absen['ID KARYAWAN'].astype(str).str.zfill(8)
        absen['TGL_DT'] = pd.to_datetime(absen['TANGGAL'], format='%Y-%m-%d', errors='coerce')
    return db, absen
db_df, absen_df = load_data()

def generate_xlsx(id_kar, tgl_awal, tgl_akhir):
    df = absen_df[(absen_df['ID KARYAWAN']==id_kar) & (absen_df['TGL_DT']>=pd.to_datetime(tgl_awal)) & (absen_df['TGL_DT']<=pd.to_datetime(tgl_akhir))].copy()
    if df.empty:
        return None

    hadir = len(df[df['STATUS']=='H'])
    alfa = len(df[df['STATUS']=='A'])
    telat = 0 # logic telat mu
    df['JAM_LEMBUR_F'] = pd.to_numeric(df['JAM LEMBUR'], errors='coerce').fillna(0)
    total_lembur = df['JAM_LEMBUR_F'].sum()
    jml_shift = len(df[(df['STATUS']=='H') & (df['SHIFT'].str.contains('S2|S3', na=False))])
    hari_lembur = len(df[df['JAM_LEMBUR_F']>0])

    gaji_pokok = 5252909
    premi_hadir = 50000 if alfa==0 else 0
    uang_makan = hadir * 9500
    uang_transport = hadir * 0
    uang_lembur = total_lembur * 30000
    uang_shift = jml_shift * 2187
    uang_makan_lembur = hari_lembur * 9500
    tunj_loyalitas = 3500

    jkk = round(gaji_pokok * 0.0024)
    jkm = round(gaji_pokok * 0.0030)
    jht_perusahaan = round(gaji_pokok * 0.037)
    jp_perusahaan = round(gaji_pokok * 0.02)
    bpjs_kes_perusahaan = round(gaji_pokok * 0.04)

    total_pendapatan = gaji_pokok + premi_hadir + uang_makan + uang_transport + uang_lembur + uang_shift + uang_makan_lembur + tunj_loyalitas + jkk + jkm + jht_perusahaan + jp_perusahaan + bpjs_kes_perusahaan

    pot_jkk=jkk; pot_jkm=jkm; pot_jht_perusahaan=jht_perusahaan; pot_jp_perusahaan=jp_perusahaan; pot_bpjs_kes_perusahaan=bpjs_kes_perusahaan
    pot_jht_tk=round(gaji_pokok*0.02); pot_jp_tk=round(gaji_pokok*0.01); pot_bpjs_kes_karyawan=round(gaji_pokok*0.01)
    total_potongan = pot_jkk+pot_jkm+pot_jht_perusahaan+pot_jp_perusahaan+pot_bpjs_kes_perusahaan+pot_jht_tk+pot_jp_tk+pot_bpjs_kes_karyawan
    total_gaji = total_pendapatan - total_potongan

    wb = Workbook()
    ws = wb.active
    ws.title = "DATA GAJI"
    # Style
    fill_green = PatternFill(start_color="0F4C3A", end_color="0F4C3A", fill_type="solid")
    font_white = Font(color="FFFFFF", bold=True)
    font_yellow = Font(color="FFFF00", bold=True)
    font_green = Font(color="00FF00")
    font_red = Font(color="FF0000", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["PENDAPATAN (A)", "KET (B)", "JUMLAH (C)", "POTONGAN (D)", "JUMLAH (E)"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = fill_green
        c.font = font_white
        c.alignment = center

    ws.cell(row=2, column=1, value="PENDAPATAN").font = font_yellow
    ws.cell(row=2, column=4, value="POTONGAN").font = font_red

    data_rows = [
        ["Gaji", "", gaji_pokok, "JKK (0.24%)", jkk],
        ["Premi Hadir", f"{telat} Telat", premi_hadir, "JKM (0.30%)", jkm],
        ["Uang Makan", f"{hadir} Hari x 9500", uang_makan, "JHT Perusahaan (3.7%)", jht_perusahaan],
        ["Uang Transport", f"{hadir} Hari x 0", uang_transport, "JP Perusahaan (2%)", jp_perusahaan],
        ["Uang Lembur", f"{total_lembur} Jam x 30000", uang_lembur, "BPJS Kes Perusahaan (4%)", bpjs_kes_perusahaan],
        ["Uang Shift", f"{jml_shift} Hari x 2187", uang_shift, "JHT TK (2%)", pot_jht_tk],
        ["Uang Makan Lembur", f"{hari_lembur} Hari x 9500", uang_makan_lembur, "JP TK (1%)", pot_jp_tk],
        ["Tunjangan Loyalitas", "", tunj_loyalitas, "BPJS Kes Karyawan (1%)", pot_bpjs_kes_karyawan],
        ["JKK (0.24%)", "", jkk, "", ""],
        ["JKM (0.30%)", "", jkm, "", ""],
        ["JHT Perusahaan (3.7%)", "", jht_perusahaan, "", ""],
        ["JP Perusahaan (2%)", "", jp_perusahaan, "", ""],
        ["BPJS Kes Perusahaan (4%)", "", bpjs_kes_perusahaan, "", ""],
        ["", "", "", "", ""],
        ["TOTAL PENDAPATAN", "", total_pendapatan, "TOTAL POTONGAN", total_potongan],
        ["", "", "", "", ""],
        ["TOTAL GAJI", "", total_gaji, "", ""],
    ]
    for r, row in enumerate(data_rows, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Sheet 2 REKAP ABSENSI
    ws2 = wb.create_sheet("REKAP ABSENSI PERIODE")
    ws2.append(['ID KARYAWAN','NAMA KARYAWAN','TANGGAL','JAM MASUK','JAM PULANG','JAM KERJA','JAM LEMBUR','LEMBUR 1.5','LEMBUR 2.0','SHIFT','KETERANGAN','STATUS','UANG SHIFT'])
    for _, row in df.iterrows():
        ws2.append([row['ID KARYAWAN'], row['NAMA KARYAWAN'], row['TANGGAL'], row['JAM MASUK'], row['JAM PULANG'], row['JAM KERJA'], row['JAM LEMBUR'], row['LEMBUR 1.5'], row['LEMBUR 2.0'], row['SHIFT'], row['KETERANGAN'], row['STATUS'], row['UANG SHIFT']])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio, total_gaji, hadir, total_lembur, jml_shift

# UI
mode_g = st.radio("Mode", ["21-20 Payroll","Bulan Kalender","Custom"], horizontal=True)
c1,c2 = st.columns(2)
with c1: bulan_g = st.selectbox("Bulan", list(range(1,13)), index=8)
with c2: tahun_g = st.number_input("Tahun", 2020, 2030, 2026)

if mode_g=="Custom":
    cc1,cc2 = st.columns(2)
    with cc1: awal_g = st.date_input("Dari", date(tahun_g,bulan_g,1))
    with cc2: akhir_g = st.date_input("Sampai", date(tahun_g,bulan_g,20))
else:
    if mode_g=="21-20 Payroll":
        if bulan_g==1:
            awal_g=date(tahun_g-1,12,21)
            akhir_g=date(tahun_g,1,20)
        else:
            awal_g=date(tahun_g,bulan_g-1,21)
            akhir_g=date(tahun_g,bulan_g,20)
    else:
        awal_g=date(tahun_g,bulan_g,1)
        akhir_g=date(tahun_g,bulan_g,calendar.monthrange(tahun_g,bulan_g)[1])

st.info(f"Periode: {awal_g} - {akhir_g}")

id_gaji = st.selectbox("ID Karyawan", db_df['ID KARYAWAN'].tolist() if not db_df.empty else ["01213027"])

if st.button("HITUNG & GENERATE XLSX", type="primary", use_container_width=True):
    bio, total_gaji, hadir, total_lembur, jml_shift = generate_xlsx(id_gaji, awal_g, akhir_g)
    if bio is None:
        st.error("Data kosong")
    else:
        st.success(f"✅ Hadir: {hadir} | Lembur: {total_lembur} | Shift: {jml_shift} | TOTAL: Rp {total_gaji:,}")
        st.download_button(
            label="📥 DOWNLOAD XLSX (Format REKAP)",
            data=bio,
            file_name=f"GAJI_{id_gaji}_{awal_g}_{akhir_g}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
)
